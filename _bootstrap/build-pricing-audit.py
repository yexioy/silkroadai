#!/usr/bin/env python3
"""
W7 D1 → W7 D2 — pricing audit xlsx generator.

Inputs:
  /tmp/newapi-pricing.json         — `/api/pricing` dump from new-api admin
  /tmp/newapi-channel-models.tsv   — channel_id, name, type, models CSV
  Hard-coded SF wholesale + comparison data below (from web fetches; cited)

Output (set via --w7d2 flag):
  /Users/mac/Documents/silk road ai/docs/W7-D1-pricing-audit.xlsx (D1 default)
  /Users/mac/Documents/silk road ai/docs/W7-D2-pricing-audit.xlsx (post-cutover)

Constants (W7 D2 cutover flipped these — pass --w7d2 to use post-cutover values):
  QPU_PER_USD: 500_000 (D1) → 1_000_000 (D2)
  USD_TO_CNY:  7.2 (D1)      → 7.0 (D2 fixed)
"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

W7D2_MODE = '--w7d2' in sys.argv
USD_TO_CNY = 7.0 if W7D2_MODE else 7.2
# After QPU=1M (W7 D2 cutover), USD/1M_tokens = mr × 1. Pre-cutover at QPU=500K it was mr × 2.
USD_PER_MR_PER_1M = 1.0 if W7D2_MODE else 2.0

# =============================================================================
# Load data
# =============================================================================
pricing = json.load(open('/tmp/newapi-pricing.json'))['data']
by_name = {r['model_name']: r for r in pricing}

# Channel → model list (id, name, type, models). TSV from psql -t -A,
# columns are tab-separated; one row per channel.
ch_data = {}
for line in open('/tmp/newapi-channel-models.tsv'):
    if not line.strip():
        continue
    parts = line.rstrip('\n').split('\t')
    if len(parts) < 4 or not parts[0].strip().isdigit():
        continue
    cid, cname, ctype, models = parts[0].strip(), parts[1].strip(), parts[2].strip(), parts[3].strip()
    ch_data[int(cid)] = {
        'name': cname,
        'type': int(ctype),
        'models': [m.strip() for m in models.split(',') if m.strip()],
    }

# Reverse: model → channel(s).  W3 D2.5 added W1 short-name aliases under
# the same channel as canonical names, so a model may appear ONCE per
# channel only.
def channel_of(mname):
    matches = []
    for cid, info in ch_data.items():
        if mname in info['models']:
            matches.append((cid, info['name'], info['type']))
    return matches

# =============================================================================
# SF WHOLESALE — from https://www.siliconflow.com/pricing (web-fetched 2026-05-06)
# Pro tier where listed; standard otherwise. Numbers in ¥ per 1M tokens.
# =============================================================================
SF_WHOLESALE = {
    # canonical name → (input_cny_1m, output_cny_1m, source_note)
    'deepseek-ai/DeepSeek-V4-Flash':       (0.14, 0.28, 'siliconflow.com/pricing'),
    'Pro/deepseek-ai/DeepSeek-V3.2':       (0.27, 0.42, 'siliconflow.com/pricing'),
    'deepseek-ai/DeepSeek-V3.2':           (0.27, 0.42, 'siliconflow.com/pricing'),
    'Pro/deepseek-ai/DeepSeek-V3.1-Terminus': (0.27, 1.0, 'siliconflow.com/pricing'),
    'deepseek-ai/DeepSeek-V3.1-Terminus':  (0.27, 1.0, 'siliconflow.com/pricing'),
    'Pro/deepseek-ai/DeepSeek-V3':         (2.0, 8.0, 'web-search early-2025 SF Pro tier'),
    'deepseek-ai/DeepSeek-V3':             (1.0, 2.0, 'web-search early-2025 SF standard tier'),
    'Pro/deepseek-ai/DeepSeek-R1':         (4.0, 16.0, 'web-search early-2025 SF Pro tier'),
    'deepseek-ai/DeepSeek-R1':             (4.0, 16.0, 'web-search early-2025 SF (R1 standard = Pro)'),
    'Pro/zai-org/GLM-4.7':                 (0.42, 2.2, 'siliconflow.com/pricing'),
    'zai-org/GLM-4.6':                     (0.39, 1.9, 'siliconflow.com/pricing'),
    'Pro/moonshotai/Kimi-K2-Instruct-0905':(0.4, 2.0, 'siliconflow.com/pricing'),
    'moonshotai/Kimi-K2-Instruct-0905':    (0.4, 2.0, 'siliconflow.com/pricing'),
    'Pro/moonshotai/Kimi-K2-Thinking':     (4.0, 16.0, 'web-search 2026 SF (estimate from official Kimi pricing)'),
    'moonshotai/Kimi-K2-Thinking':         (4.0, 16.0, 'web-search 2026 SF (estimate)'),
    'tencent/Hunyuan-A13B-Instruct':       (0.14, 0.57, 'siliconflow.com/pricing'),
    'MiniMaxAI/MiniMax-M2.5':              (0.3, 1.2, 'siliconflow.com/pricing'),
    'Pro/MiniMaxAI/MiniMax-M2.5':          (0.3, 1.2, 'siliconflow.com/pricing'),
    'Qwen/Qwen3-VL-32B-Instruct':          (0.2, 0.6, 'siliconflow.com/pricing'),
    'Qwen/Qwen3-VL-8B-Instruct':           (0.18, 0.68, 'siliconflow.com/pricing'),
    'BAAI/bge-m3':                         (0.0, 0.0, 'SF promo: free / 限时免费 (verify before launch)'),
    'Pro/BAAI/bge-m3':                     (0.0, 0.0, 'SF promo: free / 限时免费 (verify before launch)'),
}

# =============================================================================
# UPSTREAM RETAIL (USD/1M) — for "vs Direct" sheet
# Sources: openai.com/api/pricing, anthropic.com/api/pricing
# =============================================================================
UPSTREAM_RETAIL_USD = {
    # OpenAI (as of 2026-05; 5.x family is post-launch est.)
    'gpt-4o':                   (2.50, 10.00, 'OpenAI direct'),
    'gpt-4o-mini':              (0.15, 0.60, 'OpenAI direct'),
    'gpt-4-turbo':              (10.0, 30.0, 'OpenAI direct'),
    'gpt-3.5-turbo':            (0.50, 1.50, 'OpenAI direct'),
    'gpt-4.1':                  (2.00, 8.00, 'OpenAI direct'),
    'gpt-4.1-mini':             (0.40, 1.60, 'OpenAI direct'),
    'gpt-4.1-nano':             (0.10, 0.40, 'OpenAI direct'),
    'o1':                       (15.0, 60.0, 'OpenAI direct'),
    'o1-mini':                  (1.10, 4.40, 'OpenAI direct'),
    'o3':                       (10.0, 40.0, 'OpenAI direct'),
    'o3-mini':                  (1.10, 4.40, 'OpenAI direct'),
    'o4-mini':                  (1.10, 4.40, 'OpenAI direct'),
    'gpt-5':                    (2.00, 8.00, 'OpenAI direct (est.)'),
    'gpt-5-mini':               (0.40, 1.60, 'OpenAI direct (est.)'),
    'gpt-5-nano':               (0.10, 0.40, 'OpenAI direct (est.)'),
    # Anthropic
    'claude-opus-4-1-20250805': (15.0, 75.0, 'Anthropic direct'),
    'claude-opus-4-20250514':   (15.0, 75.0, 'Anthropic direct'),
    'claude-opus-4-5-20251101': (15.0, 75.0, 'Anthropic direct'),
    'claude-opus-4-7':          (15.0, 75.0, 'Anthropic direct (est. — Opus 4.x line)'),
    'claude-sonnet-4-20250514': (3.0, 15.0, 'Anthropic direct'),
    'claude-sonnet-4-5-20250929':(3.0, 15.0, 'Anthropic direct'),
    'claude-sonnet-4-6':        (3.0, 15.0, 'Anthropic direct (est.)'),
    'claude-3-5-sonnet-20241022':(3.0, 15.0, 'Anthropic direct'),
    'claude-3-5-haiku-20241022':(0.80, 4.0, 'Anthropic direct'),
    'claude-haiku-4-5-20251001':(1.0, 5.0, 'Anthropic direct (est.)'),
    'claude-3-opus-20240229':   (15.0, 75.0, 'Anthropic direct'),
    # Embeddings
    'text-embedding-3-large':   (0.13, 0.13, 'OpenAI direct'),
    'text-embedding-3-small':   (0.02, 0.02, 'OpenAI direct'),
    'text-embedding-ada-002':   (0.10, 0.10, 'OpenAI direct'),
    # Audio (USD/1M tokens equiv; for whisper actual is $0.006/min — left as-is for new-api accounting)
    'whisper-1':                (0.0, 0.0, 'OpenAI direct $0.006/min — token model not directly comparable'),
}

# =============================================================================
# Customer-side price (from new-api ratios)
# =============================================================================
DEFAULT_FALLBACK_RATIO = 37.5  # 261/379 models in prod sit here — almost certainly the
                                # new-api default fallback for unmapped model_ratio entries.

def customer_prices(model_name):
    r = by_name.get(model_name)
    if not r:
        return None
    is_default_fallback = (r['model_ratio'] == DEFAULT_FALLBACK_RATIO and r.get('completion_ratio', 1) == 1)
    if r.get('quota_type', 0) == 1 and r.get('model_price', 0) > 0:
        # fixed per-call price — return as a marker (USD per call)
        return {
            'mode': 'per_call',
            'price_usd_per_call': r['model_price'],
            'in_usd_1m': None,
            'out_usd_1m': None,
            'in_cny_1m': None,
            'out_cny_1m': None,
            'mr': r['model_ratio'],
            'cr': r.get('completion_ratio', 1),
            'is_default_fallback': is_default_fallback,
        }
    in_usd = r['model_ratio'] * USD_PER_MR_PER_1M
    out_usd = r['model_ratio'] * r.get('completion_ratio', 1) * USD_PER_MR_PER_1M
    return {
        'mode': 'per_token',
        'in_usd_1m': in_usd,
        'out_usd_1m': out_usd,
        'in_cny_1m': in_usd * USD_TO_CNY,
        'out_cny_1m': out_usd * USD_TO_CNY,
        'mr': r['model_ratio'],
        'cr': r.get('completion_ratio', 1),
        'is_default_fallback': is_default_fallback,
    }

# =============================================================================
# Master rows — pick representative subset (alias and canonical de-duped)
# =============================================================================
if W7D2_MODE:
    # Post-cutover whitelist (operator-confirmed via project_silkroadai_pricing_strategy.md):
    # 6 Anthropic + 8 OpenAI + 22 SF = 36 SKU. No legacy/dated variants.
    PRIORITY_MODELS = [
        # SF (22 — cost-plus pricing, 0 promo)
        'BAAI/bge-m3',
        'Pro/BAAI/bge-m3',
        'MiniMaxAI/MiniMax-M2.5',
        'Pro/MiniMaxAI/MiniMax-M2.5',
        'deepseek-ai/DeepSeek-V4-Flash',
        'tencent/Hunyuan-A13B-Instruct',
        'Qwen/Qwen3-VL-8B-Instruct',
        'Qwen/Qwen3-VL-32B-Instruct',
        'Pro/deepseek-ai/DeepSeek-V3.1-Terminus',
        'Pro/deepseek-ai/DeepSeek-V3.2',
        'deepseek-ai/DeepSeek-V3.1-Terminus',
        'deepseek-ai/DeepSeek-V3.2',
        'zai-org/GLM-4.6',
        'Pro/moonshotai/Kimi-K2-Instruct-0905',
        'moonshotai/Kimi-K2-Instruct-0905',
        'Pro/zai-org/GLM-4.7',
        'deepseek-ai/DeepSeek-V3',
        'Pro/deepseek-ai/DeepSeek-V3',
        'Pro/deepseek-ai/DeepSeek-R1',
        'Pro/moonshotai/Kimi-K2-Thinking',
        'deepseek-ai/DeepSeek-R1',
        'moonshotai/Kimi-K2-Thinking',
        # sub2api Anthropic (6 — 50% promo through 2026-06-09)
        'claude-opus-4-7',
        'claude-opus-4-6',
        'claude-opus-4-5',
        'claude-sonnet-4-6',
        'claude-sonnet-4-5',
        'claude-haiku-4-5',
        # sub2api-openai (8 — 50% promo through 2026-06-09)
        'gpt-5.2',
        'gpt-5.3-codex',
        'gpt-5.4',
        'gpt-5.4-mini',
        'gpt-5.5',
        'gpt-4o-audio-preview',
        'gpt-4o-realtime-preview',
        'gpt-image-1.5',
    ]
else:
    PRIORITY_MODELS = [
    # SiliconFlow (canonical only — short-aliases are echos under same channel)
    'deepseek-ai/DeepSeek-V4-Flash',
    'Pro/deepseek-ai/DeepSeek-V3.2',
    'Pro/deepseek-ai/DeepSeek-V3.1-Terminus',
    'deepseek-ai/DeepSeek-R1',
    'deepseek-ai/DeepSeek-V3',
    'Qwen/Qwen3-235B-A22B-Instruct-2507',
    'Qwen/Qwen2.5-72B-Instruct',
    'Qwen/Qwen3-VL-32B-Instruct',
    'Qwen/Qwen3-VL-8B-Instruct',
    'Pro/zai-org/GLM-4.7',
    'zai-org/GLM-4.6',
    'zai-org/GLM-4.5-Air',
    'Pro/moonshotai/Kimi-K2-Instruct-0905',
    'Pro/moonshotai/Kimi-K2-Thinking',
    'tencent/Hunyuan-A13B-Instruct',
    'tencent/Hunyuan-MT-7B',
    'Pro/MiniMaxAI/MiniMax-M2.5',
    'BAAI/bge-m3',
    'Pro/BAAI/bge-m3',
    'Qwen/Qwen3-Embedding-8B',
    # sub2api Anthropic
    'claude-opus-4-7',
    'claude-opus-4-7-thinking',
    'claude-sonnet-4-6',
    'claude-opus-4-6',
    'claude-opus-4-5-20251101',
    'claude-sonnet-4-5-20250929',
    'claude-sonnet-4-5-20250929-thinking',
    'claude-sonnet-4-20250514',
    'claude-opus-4-20250514',
    'claude-opus-4-1-20250805',
    'claude-3-5-sonnet-20241022',
    'claude-3-5-haiku-20241022',
    'claude-haiku-4-5-20251001',
    'claude-3-opus-20240229',
    # sub2api OpenAI
    'gpt-4o',
    'gpt-4o-mini',
    'gpt-4-turbo',
    'gpt-3.5-turbo',
    'gpt-4.1',
    'gpt-4.1-mini',
    'gpt-4.1-nano',
    'gpt-5',
    'gpt-5-mini',
    'gpt-5-codex',
    'gpt-5-pro',
    'gpt-5.4',
    'gpt-5.4-pro',
    'o1',
    'o1-mini',
    'o3',
    'o3-mini',
    'o4-mini',
    'whisper-1',
    'tts-1',
    'tts-1-hd',
    'gpt-4o-mini-tts',
    'gpt-4o-transcribe',
    'text-embedding-3-large',
    'text-embedding-3-small',
    'text-embedding-ada-002',
    'dall-e-3',
    'dall-e-2',
    'sora-2',
]

CHANNEL_LABEL = {
    1: 'siliconflow',
    2: 'sub2api (Anthropic)',
    3: 'sub2api-openai',
}

def build_master_rows():
    rows = []
    for mname in PRIORITY_MODELS:
        p = customer_prices(mname)
        if not p:
            continue
        chs = channel_of(mname)
        if not chs:
            continue
        cid, cname, ctype = chs[0]
        chlabel = CHANNEL_LABEL.get(cid, cname)
        ws = SF_WHOLESALE.get(mname)
        if ws:
            ws_in, ws_out, ws_note = ws
        elif cid == 1:
            ws_in, ws_out, ws_note = (None, None, 'TBD: SF page did not list this model — fetch from cloud.siliconflow.cn before launch')
        else:
            ws_in, ws_out, ws_note = (None, None, '运营者填:实际 sub2api 月度账单 / 1M token 平均成本')
        # margin
        if p['mode'] == 'per_token' and ws_in is not None and ws_out is not None and p['in_cny_1m']:
            in_margin = (p['in_cny_1m'] - ws_in) / p['in_cny_1m'] if p['in_cny_1m'] else None
            # weight margins 1:3 (input vs output) since in real chat usage,
            # output dominates token volume; this gives a "blended" feel.
            blended_customer = p['in_cny_1m'] + 3 * p['out_cny_1m']
            blended_wholesale = ws_in + 3 * ws_out
            margin = (blended_customer - blended_wholesale) / blended_customer if blended_customer else None
        else:
            in_margin = None
            margin = None
        rows.append({
            'channel': chlabel,
            'cid': cid,
            'model': mname,
            'mode': p['mode'],
            'mr': p['mr'],
            'cr': p['cr'],
            'in_usd': p['in_usd_1m'],
            'out_usd': p['out_usd_1m'],
            'in_cny': p['in_cny_1m'],
            'out_cny': p['out_cny_1m'],
            'ws_in': ws_in,
            'ws_out': ws_out,
            'ws_note': ws_note,
            'margin': margin,
            'is_default_fallback': p['is_default_fallback'],
        })
    # Sort by channel then by customer in_cny desc (most expensive at top)
    rows.sort(key=lambda r: (r['cid'], -(r['in_cny'] or 0)))
    return rows

rows = build_master_rows()
print(f"master rows: {len(rows)}")

# =============================================================================
# Build xlsx
# =============================================================================
wb = Workbook()

# Styling
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill('solid', fgColor='0A1535')
THIN = Side(border_style='thin', color='D0D5E0')
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)

CHANNEL_FILLS = {
    'siliconflow':         PatternFill('solid', fgColor='E8F5FF'),
    'sub2api (Anthropic)': PatternFill('solid', fgColor='FFF4E0'),
    'sub2api-openai':      PatternFill('solid', fgColor='F0E8FF'),
}

# margin band fills
RED    = PatternFill('solid', fgColor='FDECEA')
YELLOW = PatternFill('solid', fgColor='FFF8E1')
GREEN  = PatternFill('solid', fgColor='E8F5E9')
BLUE   = PatternFill('solid', fgColor='E3F2FD')

def margin_fill(margin):
    if margin is None:
        return None
    # W7 D2 thresholds: SF cost-plus formula intentionally targets ~17%
    # blended margin ("+20% markup over wholesale" per
    # project_silkroadai_pricing_strategy.md). Anything ≥ 15% is by-design
    # healthy GREEN. RED is now reserved for genuine concern (< 10%).
    if margin < 0.10:
        return RED
    if margin < 0.15:
        return YELLOW
    if margin < 0.50:
        return GREEN
    return BLUE

def margin_label(margin):
    if margin is None:
        return 'TBD'
    if margin < 0:
        return 'LOSS (亏本)'
    if margin < 0.10:
        return 'red (<10%)'
    if margin < 0.15:
        return 'yellow (10-15%)'
    if margin < 0.50:
        return 'green (15-50%)'
    return 'blue (>50%)'

# -----------------------------------------------------------------------------
# Sheet 1 — Pricing Master
# -----------------------------------------------------------------------------
ws1 = wb.active
ws1.title = 'Pricing Master'
HEADERS_1 = [
    'Channel', 'Model', 'Type', 'model_ratio', 'completion_ratio',
    'Default Fallback?',
    'Input USD/1M', 'Output USD/1M',
    'Customer ¥/1M In', 'Customer ¥/1M Out',
    'Wholesale ¥/1M In', 'Wholesale ¥/1M Out',
    'Margin %', 'Notes',
]
for i, h in enumerate(HEADERS_1, 1):
    c = ws1.cell(row=1, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER
ws1.row_dimensions[1].height = 32

FALLBACK_FILL = PatternFill('solid', fgColor='FFE0E0')  # pinkish — flag misconfig
for ri, r in enumerate(rows, 2):
    ws1.cell(row=ri, column=1, value=r['channel'])
    ws1.cell(row=ri, column=2, value=r['model'])
    ws1.cell(row=ri, column=3, value=r['mode'])
    ws1.cell(row=ri, column=4, value=r['mr'])
    ws1.cell(row=ri, column=5, value=r['cr'])
    fb = ws1.cell(row=ri, column=6, value='⚠️ YES (37.5 default)' if r.get('is_default_fallback') else '')
    if r.get('is_default_fallback'):
        fb.fill = FALLBACK_FILL
    ws1.cell(row=ri, column=7, value=r['in_usd'])
    ws1.cell(row=ri, column=8, value=r['out_usd'])
    ws1.cell(row=ri, column=9, value=r['in_cny'])
    ws1.cell(row=ri, column=10, value=r['out_cny'])
    ws1.cell(row=ri, column=11, value=r['ws_in'])
    ws1.cell(row=ri, column=12, value=r['ws_out'])
    # Live formula for margin: blended (in + 3*out) so when wholesale is
    # filled in later by operator, margin auto-updates. Columns shifted
    # by +1 because of the new "Default Fallback?" column at col 6.
    margin_formula = (
        f'=IF(OR(ISBLANK(K{ri}),ISBLANK(L{ri})),"TBD",'
        f'((I{ri}+3*J{ri})-(K{ri}+3*L{ri}))/(I{ri}+3*J{ri}))'
    )
    mc = ws1.cell(row=ri, column=13, value=margin_formula)
    mc.number_format = '0.0%'
    if r['margin'] is not None:
        mf = margin_fill(r['margin'])
        if mf:
            mc.fill = mf
    ws1.cell(row=ri, column=14, value=r['ws_note'] or '')
    # Borders all cells
    for col in range(1, 15):
        ws1.cell(row=ri, column=col).border = BORDER

    # Number formats (shifted)
    for col, fmt in [(4, '0.0000'), (5, '0.00'), (7, '$0.0000'), (8, '$0.0000'),
                     (9, '¥0.00'), (10, '¥0.00'), (11, '¥0.00'), (12, '¥0.00')]:
        ws1.cell(row=ri, column=col).number_format = fmt

# Column widths (14 cols now)
WIDTHS_1 = [22, 42, 11, 12, 12, 22, 13, 13, 16, 17, 17, 18, 11, 60]
for i, w in enumerate(WIDTHS_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = 'A2'

# -----------------------------------------------------------------------------
# Sheet 2 — Margin Alerts
# -----------------------------------------------------------------------------
ws2 = wb.create_sheet('Margin Alerts')
HEADERS_2 = ['Severity', 'Channel', 'Model', 'Customer ¥/1M In', 'Customer ¥/1M Out',
             'Wholesale ¥/1M In', 'Wholesale ¥/1M Out', 'Blended Margin %', 'Action 建议']
for i, h in enumerate(HEADERS_2, 1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER
ws2.row_dimensions[1].height = 30

# Bands buckets — W7 D2 thresholds aligned with margin_fill() / margin_label():
# RED   < 10%  (genuine concern; SF formula yields 17% by design so should NOT hit)
# YELLOW 10-15% (thin but acceptable transitional)
# GREEN  15-50% (healthy; SF cost-plus & sub2api promo land here)
# BLUE   > 50%  (over-priced — flag for "do customers find replacements?")
ALERT_ROWS = []
for r in rows:
    if r['margin'] is None:
        ALERT_ROWS.append(('TBD', r))
    elif r['margin'] < 0:
        ALERT_ROWS.append(('LOSS', r))
    elif r['margin'] < 0.10:
        ALERT_ROWS.append(('RED', r))
    elif r['margin'] < 0.15:
        ALERT_ROWS.append(('YELLOW', r))
    elif r['margin'] >= 0.50:
        ALERT_ROWS.append(('BLUE', r))
    # GREEN models omitted — only show needs-attention rows
SEVERITY_ORDER = {'LOSS': 0, 'RED': 1, 'YELLOW': 2, 'BLUE': 3, 'TBD': 4}
ALERT_ROWS.sort(key=lambda x: (SEVERITY_ORDER[x[0]], -(x[1].get('in_cny') or 0)))

ACTION_FOR = {
    'LOSS':   '⚠️ 亏本卖,立即调价或下架。',
    'RED':    '🟥 毛利<10%,极薄;调价或与上游议价。',
    'YELLOW': '🟨 毛利 10-15%,薄但 W7 战略可接受。',
    'BLUE':   '🟦 毛利>50%,客户可能找替代,考虑微降以促转化。',
    'TBD':    '⏳ 无 wholesale 数据(sub2api 渠道),margin 计算 N/A。',
}
SEVERITY_FILL = {'LOSS': RED, 'RED': RED, 'YELLOW': YELLOW, 'BLUE': BLUE, 'TBD': PatternFill('solid', fgColor='F5F5F5')}

for ri, (sev, r) in enumerate(ALERT_ROWS, 2):
    ws2.cell(row=ri, column=1, value=sev)
    ws2.cell(row=ri, column=2, value=r['channel'])
    ws2.cell(row=ri, column=3, value=r['model'])
    ws2.cell(row=ri, column=4, value=r['in_cny'])
    ws2.cell(row=ri, column=5, value=r['out_cny'])
    ws2.cell(row=ri, column=6, value=r['ws_in'])
    ws2.cell(row=ri, column=7, value=r['ws_out'])
    if r['margin'] is not None:
        mc = ws2.cell(row=ri, column=8, value=r['margin'])
        mc.number_format = '0.0%'
    else:
        ws2.cell(row=ri, column=8, value='TBD')
    ws2.cell(row=ri, column=9, value=ACTION_FOR[sev])
    f = SEVERITY_FILL[sev]
    for col in range(1, 10):
        cell = ws2.cell(row=ri, column=col)
        cell.fill = f
        cell.border = BORDER
        if col in (4, 5, 6, 7):
            cell.number_format = '¥0.00'

WIDTHS_2 = [10, 22, 42, 16, 17, 17, 17, 16, 70]
for i, w in enumerate(WIDTHS_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'

# -----------------------------------------------------------------------------
# Sheet 3 — vs Direct OpenAI / Anthropic
# -----------------------------------------------------------------------------
ws3 = wb.create_sheet('vs Direct')
HEADERS_3 = ['Model', 'Channel', 'Our ¥/1M In', 'Our ¥/1M Out',
             'Direct ¥/1M In', 'Direct ¥/1M Out', 'In Ratio (我们/上游)', 'Out Ratio', 'Risk', 'Notes']
for i, h in enumerate(HEADERS_3, 1):
    c = ws3.cell(row=1, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = BORDER
ws3.row_dimensions[1].height = 30

vs_rows = []
for r in rows:
    upstream = UPSTREAM_RETAIL_USD.get(r['model'])
    if not upstream or r['mode'] != 'per_token':
        continue
    up_in_usd, up_out_usd, up_note = upstream
    up_in_cny = up_in_usd * USD_TO_CNY
    up_out_cny = up_out_usd * USD_TO_CNY
    in_ratio = (r['in_cny'] / up_in_cny) if up_in_cny else None
    out_ratio = (r['out_cny'] / up_out_cny) if up_out_cny else None
    # risk level on the higher ratio (worst-case visibility for sophisticated users)
    worst_ratio = max([x for x in (in_ratio, out_ratio) if x is not None], default=None)
    if worst_ratio is None:
        risk = 'N/A'
    elif worst_ratio > 2:
        risk = 'RED — 用户大概率绕过我们直购上游'
    elif worst_ratio > 1.5:
        risk = 'YELLOW — 价差太大,聪明用户会比价'
    elif worst_ratio > 1.0:
        risk = 'GREEN — 合理溢价(便利性 + 中国大陆访问)'
    else:
        risk = 'BLUE — 我们价格 ≤ 上游(中国市场可消化的负毛利或亏本卖)'
    vs_rows.append({
        'model': r['model'], 'channel': r['channel'],
        'our_in': r['in_cny'], 'our_out': r['out_cny'],
        'up_in': up_in_cny, 'up_out': up_out_cny,
        'in_ratio': in_ratio, 'out_ratio': out_ratio,
        'risk': risk, 'note': up_note,
    })

vs_rows.sort(key=lambda r: -(max(filter(lambda x: x is not None, [r['in_ratio'], r['out_ratio']]), default=0) or 0))

VS_FILL = {
    'RED': RED, 'YELLOW': YELLOW, 'GREEN': GREEN, 'BLUE': BLUE, 'N/A': PatternFill('solid', fgColor='F5F5F5'),
}
for ri, r in enumerate(vs_rows, 2):
    ws3.cell(row=ri, column=1, value=r['model'])
    ws3.cell(row=ri, column=2, value=r['channel'])
    ws3.cell(row=ri, column=3, value=r['our_in'])
    ws3.cell(row=ri, column=4, value=r['our_out'])
    ws3.cell(row=ri, column=5, value=r['up_in'])
    ws3.cell(row=ri, column=6, value=r['up_out'])
    ws3.cell(row=ri, column=7, value=r['in_ratio'])
    ws3.cell(row=ri, column=8, value=r['out_ratio'])
    ws3.cell(row=ri, column=9, value=r['risk'].split(' — ')[0])
    ws3.cell(row=ri, column=10, value=r['risk'] + ' · ' + (r['note'] or ''))
    for col, fmt in [(3,'¥0.00'),(4,'¥0.00'),(5,'¥0.00'),(6,'¥0.00'),(7,'0.00x'),(8,'0.00x')]:
        ws3.cell(row=ri, column=col).number_format = fmt
    risk_key = r['risk'].split(' — ')[0]
    fill = VS_FILL.get(risk_key)
    for col in range(1, 11):
        cell = ws3.cell(row=ri, column=col)
        if fill:
            cell.fill = fill
        cell.border = BORDER

WIDTHS_3 = [38, 22, 14, 14, 14, 14, 18, 12, 8, 60]
for i, w in enumerate(WIDTHS_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = 'A2'

# -----------------------------------------------------------------------------
out_path = (
    '/Users/mac/Documents/silk road ai/docs/W7-D2-pricing-audit.xlsx'
    if W7D2_MODE
    else '/Users/mac/Documents/silk road ai/docs/W7-D1-pricing-audit.xlsx'
)
wb.save(out_path)
print(f"saved → {out_path}")
print(f"  sheet 1 rows: {len(rows)}")
print(f"  sheet 2 rows: {len(ALERT_ROWS)}")
print(f"  sheet 3 rows: {len(vs_rows)}")

# Print summary breakdown for the briefing
import collections
band_counts = collections.Counter()
for r in rows:
    if r['margin'] is None:
        band_counts['TBD'] += 1
    elif r['margin'] < 0:
        band_counts['LOSS'] += 1
    elif r['margin'] < 0.10:
        band_counts['RED'] += 1
    elif r['margin'] < 0.15:
        band_counts['YELLOW'] += 1
    elif r['margin'] < 0.50:
        band_counts['GREEN'] += 1
    else:
        band_counts['BLUE'] += 1
print(f"\nMargin distribution:")
for k in ['LOSS', 'RED', 'YELLOW', 'GREEN', 'BLUE', 'TBD']:
    print(f"  {k}: {band_counts[k]}")
